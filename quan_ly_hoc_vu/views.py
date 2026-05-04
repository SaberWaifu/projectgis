import random
import math
import os
import string
from django.core.mail import send_mail
from django.conf import settings
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.forms import AuthenticationForm
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.db.models import Count, Sum
from django.contrib.auth import update_session_auth_hash
from django.contrib.auth.forms import PasswordChangeForm
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth.models import User

from .utils import chay_tool_dieu_phoi
from .ga_solver import GeneticScheduler
# [ĐÃ SỬA] Thêm AnhMinhChung vào đây để không bị lỗi NameError
from .models import (CoSo, HocVien, LichHoc, GiaoVien, PhongHoc, LopHoc, HoaDon, KhoaHoc,
                     AnhCoSo, DanhGia, AnhMinhChung, TinTuc, AnhTinTuc, BaiGiang, VideoBaiGiang,
                     FileBaiGiang, BaiTap, NopBaiTap, ThoiGianRanh, XinNghi, LichBu, DangKyLichBu,
                     GioiThieu, AnhGioiThieu)
# ===================================================================
# HÀM KIỂM TRA QUYỀN (ROLE CHECK)
# ===================================================================
def check_admin(user):
    return user.is_authenticated and (user.is_superuser or user.is_staff)

def admin_required(view_func):
    @login_required(login_url='login')
    def _wrapped(request, *args, **kwargs):
        if not check_admin(request.user):
            return render(request, '403.html', status=403)
        return view_func(request, *args, **kwargs)
    return _wrapped

def handler403(request, exception=None):
    return render(request, '403.html', status=403)

def tinh_khoang_cach_haversine(lat1, lon1, lat2, lon2):
    try:
        lat1, lon1, lat2, lon2 = map(float, [lat1, lon1, lat2, lon2])
        R = 6371 
        dLat = math.radians(lat2 - lat1)
        dLon = math.radians(lon2 - lon1)
        a = math.sin(dLat/2)**2 + math.cos(math.radians(lat1)) * math.cos(math.radians(lat2)) * math.sin(dLon/2)**2
        c = 2 * math.asin(math.sqrt(a))
        return R * c
    except (ValueError, TypeError):
        return float('inf')

# ===================================================================
# 1. QUẢN LÝ TÀI KHOẢN (AUTHENTICATION)
# ===================================================================
def login_view(request):
    if request.user.is_authenticated:
        username_sach = request.user.username.strip()
        if check_admin(request.user):
            return redirect('thong_ke_page') 
        elif GiaoVien.objects.filter(tai_khoan__iexact=username_sach).exists(): 
            return redirect('trang_giao_vien')
        else:
            return redirect('trang_nguoi_dung') 

    if request.method == 'POST':
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            user = form.get_user()
            login(request, user)
            messages.success(request, f"Xin chào, {user.username}!")
            
            username_sach = user.username.strip()
            if check_admin(user):
                return redirect('thong_ke_page') 
            elif GiaoVien.objects.filter(tai_khoan__iexact=username_sach).exists(): 
                return redirect('trang_giao_vien')
            else:
                return redirect('trang_nguoi_dung') 
        else:
            messages.error(request, "Tên đăng nhập hoặc mật khẩu không đúng!")
    
    form = AuthenticationForm()
    return render(request, 'login.html', {'form': form})

def logout_view(request):
    logout(request)
    messages.info(request, "Đã đăng xuất thành công.")
    return redirect('trang_chu')

@login_required(login_url='login')
def xoa_tai_khoan(request):
    if request.method == 'POST':
        user = request.user
        logout(request)
        user.delete()
        messages.success(request, "Tài khoản của bạn đã được xóa thành công.")
        return redirect('trang_chu')
    return redirect('trang_nguoi_dung')

def dang_ky_tai_khoan(request):
    if request.user.is_authenticated:
        username_sach = request.user.username.strip()
        if check_admin(request.user):
            return redirect('thong_ke_page')
        elif GiaoVien.objects.filter(tai_khoan__iexact=username_sach).exists():
            return redirect('trang_giao_vien')
        else:
            return redirect('trang_nguoi_dung')

    if request.method == 'POST':
        username = request.POST.get('username')
        mat_khau = request.POST.get('password')
        xac_nhan = request.POST.get('confirm_password')

        if not username or not mat_khau:
            messages.error(request, "Vui lòng nhập đầy đủ Tên đăng nhập và Mật khẩu!")
        elif mat_khau != xac_nhan:
            messages.error(request, "Mật khẩu xác nhận không khớp!")
        elif User.objects.filter(username__iexact=username.strip()).exists():
            messages.error(request, "Tên đăng nhập này đã có người sử dụng!")
        else:
            try:
                user = User.objects.create_user(username=username.strip(), password=mat_khau)
                messages.success(request, "Tạo tài khoản thành công! Vui lòng đăng nhập.")
                return redirect('login')
            except Exception as e:
                messages.error(request, f"Có lỗi xảy ra: {str(e)}")

    return render(request, 'dang_ky_tai_khoan.html')

# ===================================================================
# TRANG DÀNH RIÊNG CHO USER THƯỜNG (HỌC VIÊN)
# ===================================================================
@login_required(login_url='login')
def trang_nguoi_dung(request):
    if check_admin(request.user):
        return redirect('thong_ke_page')

    username_sach = request.user.username.strip()
    if GiaoVien.objects.filter(tai_khoan__iexact=username_sach).exists(): 
        return redirect('trang_giao_vien')

    hv = HocVien.objects.filter(ten=request.user.username).first()
    
    if request.method == "POST":
        ten_moi = request.POST.get('ten', request.user.username)
        khoi_lop_moi = request.POST.get('khoi_lop', 12)
        dia_chi_moi = request.POST.get('dia_chi', '')  
        try:
            lat_moi = float(request.POST.get('lat', 0.0))
            lon_moi = float(request.POST.get('lon', 0.0))
        except ValueError:
            lat_moi, lon_moi = 0.0, 0.0

        if hv:
            hv.ten = ten_moi
            hv.khoi_lop = khoi_lop_moi
            hv.dia_chi = dia_chi_moi  
            
            if lat_moi != 0.0 and lon_moi != 0.0:
                hv.lat = lat_moi
                hv.lon = lon_moi
                hv.co_so_gan_nhat = None 
            hv.save()
            messages.success(request, "Đã cập nhật hồ sơ và định vị GIS thành công!")
        else:
            hv = HocVien.objects.create(ten=ten_moi, khoi_lop=khoi_lop_moi, dia_chi=dia_chi_moi, lat=lat_moi, lon=lon_moi)
            messages.success(request, "Đã khởi tạo hồ sơ không gian thành công!")
        return redirect('trang_nguoi_dung')

    ds_lop_cua_toi = LopHoc.objects.filter(hoc_viens=hv).select_related('giao_vien') if hv else []
    ds_lich_hoc = LichHoc.objects.filter(lop_hoc__in=ds_lop_cua_toi).select_related('lop_hoc', 'phong', 'lop_hoc__giao_vien').order_by('thu', 'ca') if hv else []
    tong_no = HoaDon.objects.filter(hoc_vien=hv, trang_thai__in=['CHUA', 'CHO_DUYET']).aggregate(Sum('so_tien'))['so_tien__sum'] or 0 if hv else 0

    return render(request, 'trang_nguoi_dung.html', {'user': request.user, 'hoc_vien': hv, 'ds_lop_cua_toi': ds_lop_cua_toi, 'ds_lich_hoc': ds_lich_hoc, 'tong_no': tong_no})

# ===================================================================
# TRANG DÀNH RIÊNG CHO GIÁO VIÊN
# ===================================================================
@login_required(login_url='login')
def trang_giao_vien(request):
    username_sach = request.user.username.strip()
    gv = GiaoVien.objects.filter(tai_khoan__iexact=username_sach).first() 
    if not gv:
        return redirect('trang_nguoi_dung') 

    ds_lop = LopHoc.objects.filter(giao_vien=gv).prefetch_related('hoc_viens').order_by('-id')
    tong_hoc_sinh = sum(lop.hoc_viens.count() for lop in ds_lop)
    ds_lich = LichHoc.objects.filter(lop_hoc__giao_vien=gv).select_related('lop_hoc', 'phong').order_by('thu', 'ca')
    
    return render(request, 'trang_giao_vien.html', {
        'user': request.user, 'giao_vien': gv, 'ds_lop': ds_lop,
        'ds_lich': ds_lich, 'tong_hoc_sinh': tong_hoc_sinh
    })

# ===================================================================
# 2. DASHBOARD & THỐNG KÊ (CORE)
# ===================================================================
def trang_chu(request):
    # 1. Lấy danh sách cơ sở
    danh_sach_co_so = CoSo.objects.all()
    
    # 2. Lấy danh sách đánh giá (SỬA LẠI THÀNH .filter(duyet=True))
    ds_danh_gia_home = DanhGia.objects.filter(duyet=True).select_related('co_so').order_by('-ngay_tao')[:3]
    
    # 3. Truy vấn tin tức kèm ảnh
    ds_tin_tuc_home = TinTuc.objects.prefetch_related('cac_anh').order_by('-ngay_dang')[:3]

    # 4. Giới thiệu trung tâm (lấy tối đa 3 mục đầu tiên)
    ds_gioi_thieu_home = GioiThieu.objects.prefetch_related('cac_anh').all()[:3]

    # 5. Trả dữ liệu ra giao diện
    return render(request, 'trang_chu.html', {
        'danh_sach_co_so': danh_sach_co_so,
        'ds_danh_gia_home': ds_danh_gia_home,
        'ds_tin_tuc_home': ds_tin_tuc_home,
        'ds_gioi_thieu_home': ds_gioi_thieu_home,
    })

def co_so_detail(request, cs_id):
    co_so = get_object_or_404(CoSo, id=cs_id)
    ds_lop = LopHoc.objects.filter(
        giao_vien__co_so=co_so, trang_thai__in=['MO', 'CHO']
    ).select_related('giao_vien').prefetch_related('hoc_viens').order_by('trang_thai', '-id')
    ds_giao_vien = GiaoVien.objects.filter(co_so=co_so).order_by('id')
    tong_hoc_vien = HocVien.objects.filter(co_so_gan_nhat=co_so).count()
    ds_danh_gia = co_so.danh_gia.filter(duyet=True).order_by('-ngay_tao')

    return render(request, 'co_so_detail.html', {
        'co_so': co_so, 'ds_lop': ds_lop, 'ds_giao_vien': ds_giao_vien,
        'tong_lop': ds_lop.count(), 'tong_giao_vien': ds_giao_vien.count(), 'tong_hoc_vien': tong_hoc_vien,
        'ds_danh_gia': ds_danh_gia,
    })

@admin_required 
def thong_ke_page(request):
    du_lieu = CoSo.objects.annotate(so_hv=Count('hocvien'))
    thong_ke_khoi = HocVien.objects.values('khoi_lop').annotate(count=Count('id'))
    return render(request, 'thong_ke.html', {'du_lieu': du_lieu, 'thong_ke_khoi': thong_ke_khoi})

# ===================================================================
# 3. BẢN ĐỒ SỐ & GIS
# ===================================================================
@admin_required 
def dieu_phoi_page(request):
    return render(request, 'index.html', {'hoc_vien': HocVien.objects.all(), 'co_so': CoSo.objects.all()})

@admin_required 
def kich_hoat_gis(request, hv_id):
    if chay_tool_dieu_phoi(hv_id):
        messages.success(request, "Đã chạy AI phân tích không gian OSRM thành công!")
    else:
        messages.error(request, "Lỗi phân tuyến: Không tìm thấy cơ sở hoặc tọa độ bị sai.")
    return redirect('dieu_phoi_page')

# ===================================================================
# 4. QUẢN LÝ HỌC VIÊN
# ===================================================================
def dang_ky_page(request):
    ket_qua = None 
    if request.method == "POST":
        try:
            ten = request.POST.get('ten')
            khoi = request.POST.get('khoi_lop')
            dia_chi = request.POST.get('dia_chi', '')
            lat = float(request.POST.get('lat', 0))
            lon = float(request.POST.get('lon', 0))
            
            hv = HocVien.objects.create(ten=ten, khoi_lop=khoi, dia_chi=dia_chi, lat=lat, lon=lon)
            min_dist, best_cs = float('inf'), None
            for cs in CoSo.objects.all():
                d = tinh_khoang_cach_haversine(lat, lon, cs.lat, cs.lon)
                if d < min_dist:
                    min_dist, best_cs = d, cs
            if best_cs:
                hv.co_so_gan_nhat = best_cs
                hv.save()
                ket_qua = {'hoc_vien': hv, 'co_so': best_cs, 'khoang_cach': round(min_dist, 2)}
        except Exception as e:
            messages.error(request, f"Lỗi nhập liệu: {e}")
    return render(request, 'dang_ky.html', {'ket_qua': ket_qua})

@admin_required 
def quan_ly_page(request):
    if request.method == "POST":
        try:
            student_id = request.POST.get('student_id')
            ten = request.POST.get('ten')
            khoi_lop = request.POST.get('khoi_lop')
            lat = float(request.POST.get('lat', 0))
            lon = float(request.POST.get('lon', 0))

            if student_id:
                hv = HocVien.objects.get(id=student_id)
                hv.ten, hv.khoi_lop, hv.lat, hv.lon, hv.co_so_gan_nhat = ten, khoi_lop, lat, lon, None
                hv.save()
                messages.success(request, f"Đã cập nhật hồ sơ: {hv.ten} thành công!")
            else:
                hv = HocVien.objects.create(ten=ten, khoi_lop=khoi_lop, lat=lat, lon=lon)
                messages.success(request, f"Đã lưu tọa độ cho học viên: {hv.ten}!")
        except Exception as e:
            messages.error(request, f"Lỗi: {e}")
        return redirect('quan_ly_page')
    return render(request, 'quan_ly.html', {'danh_sach': HocVien.objects.all().order_by('-id')})

@admin_required 
def xoa_hoc_vien(request, hv_id):
    get_object_or_404(HocVien, id=hv_id).delete()
    messages.success(request, "Đã xóa học viên khỏi hệ thống.")
    return redirect('quan_ly_page')

# ===================================================================
# 5. QUẢN LÝ CHI NHÁNH
# ===================================================================
@admin_required 
def chi_nhanh_page(request):
    if request.method == "POST":
        try:
            ten, dia_chi, lat, lon = request.POST.get('ten'), request.POST.get('dia_chi'), request.POST.get('lat'), request.POST.get('lon')
            mo_ta = request.POST.get('mo_ta', '')
            if not ten or not lat or not lon:
                messages.error(request, "Vui lòng nhập Tên và Tọa độ (Lat/Lon)!")
            else:
                CoSo.objects.create(ten=ten, dia_chi=dia_chi, mo_ta=mo_ta, lat=float(lat), lon=float(lon))
                messages.success(request, f"Đã khai trương cơ sở: {ten}")
        except Exception as e:
            messages.error(request, f"Lỗi hệ thống: {e}")
        return redirect('chi_nhanh_page')
    
    danh_sach = CoSo.objects.all().order_by('-id')
    stats = {'total': danh_sach.count(), 'total_students': HocVien.objects.count()}
    return render(request, 'chi_nhanh.html', {'co_so': danh_sach, 'stats': stats})

@admin_required 
def xoa_chi_nhanh(request, cs_id):
    try:
        cs = get_object_or_404(CoSo, id=cs_id)
        ten = cs.ten
        cs.delete()
        messages.success(request, f"Đã đóng cửa và xóa chi nhánh: {ten}")
    except Exception as e:
        messages.error(request, f"Lỗi khi xóa: {e}")
    return redirect('chi_nhanh_page')

def sua_chi_nhanh(request, id):
    co_so = get_object_or_404(CoSo, id=id)
    if request.method == 'POST':
        co_so.ten = request.POST.get('ten')
        co_so.dia_chi = request.POST.get('dia_chi')
        co_so.lat = request.POST.get('lat')
        co_so.lon = request.POST.get('lon')
        co_so.mo_ta = request.POST.get('mo_ta', '')
        co_so.save()
        files = request.FILES.getlist('hinh_anh')
        for i, file in enumerate(files):
            AnhCoSo.objects.create(co_so=co_so, hinh_anh=file, thu_tu=co_so.anh_co_so.count() + i)
        messages.success(request, f"Đã cập nhật thành công chi nhánh {co_so.ten}!")
        return redirect('chi_nhanh_page')
    ds_anh = co_so.anh_co_so.all()
    return render(request, 'sua_chi_nhanh.html', {'co_so': co_so, 'ds_anh': ds_anh})

@admin_required
def xoa_anh_co_so(request, anh_id):
    try:
        anh = get_object_or_404(AnhCoSo, id=anh_id)
        co_so_id = anh.co_so.id
        if anh.hinh_anh and os.path.isfile(anh.hinh_anh.path):
            os.remove(anh.hinh_anh.path)
        anh.delete()
        messages.success(request, "Đã xóa ảnh thành công!")
    except Exception as e:
        messages.error(request, f"Lỗi khi xóa ảnh: {e}")
    return redirect('sua_chi_nhanh', id=co_so_id)

# ===================================================================
# 6. QUẢN LÝ LỚP HỌC & AI XẾP LỊCH
# ===================================================================
@admin_required 
def quan_ly_lop_page(request):
    if request.method == "POST":
        try:
            class_id = request.POST.get('class_id')
            ten_lop = request.POST.get('ten_lop')
            giao_vien_id = request.POST.get('giao_vien')
            min_si_so = int(request.POST.get('min_si_so', 20))
            si_so_toi_da = int(request.POST.get('si_so_toi_da', 30))
            trang_thai = request.POST.get('trang_thai', 'CHO')

            if not ten_lop or not giao_vien_id:
                messages.error(request, "Thiếu tên lớp hoặc giáo viên!")
            else:
                if class_id: 
                    lop = LopHoc.objects.get(id=class_id)
                    lop.ten_lop = ten_lop
                    lop.giao_vien_id = giao_vien_id
                    lop.min_si_so = min_si_so
                    lop.si_so_toi_da = si_so_toi_da
                    lop.trang_thai = trang_thai
                    lop.save()
                    messages.success(request, f"Đã cập nhật thông tin lớp: {ten_lop}")
                else:
                    LopHoc.objects.create(ten_lop=ten_lop, giao_vien_id=giao_vien_id, min_si_so=min_si_so, si_so_toi_da=si_so_toi_da, trang_thai=trang_thai)
                    messages.success(request, f"Đã thiết lập lớp {ten_lop} thành công!")
        except Exception as e:
            messages.error(request, f"Lỗi hệ thống: {str(e)}")
        return redirect('quan_ly_lop_page')

    # =========================================================
    # PHẦN LỌC DỮ LIỆU VÀ PHÂN TRANG (PAGINATION)
    # =========================================================
    search_query = request.GET.get('q', '')
    ds_lop_qs = LopHoc.objects.all().select_related('giao_vien').prefetch_related('hoc_viens').order_by('-id')
    
    if search_query: 
        ds_lop_qs = ds_lop_qs.filter(ten_lop__icontains=search_query)

    stats = {
        'total_class': ds_lop_qs.count(), 
        'total_student': HocVien.objects.count(), 
        'active_class': ds_lop_qs.filter(trang_thai='MO').count()
    }

    # Phân trang: Mỗi trang sẽ hiển thị 6 lớp học (bạn có thể đổi số 6 thành 9 hoặc 12 tùy ý)
    paginator = Paginator(ds_lop_qs, 6) 
    page_number = request.GET.get('page', 1)
    ds_lop = paginator.get_page(page_number)

    return render(request, 'quan_ly_lop.html', {
        'ds_lop': ds_lop, 
        'stats': stats, 
        'ds_giao_vien': GiaoVien.objects.all(),
        'query': search_query # Biến này dùng để giữ lại từ khóa trên thanh tìm kiếm
    })

@admin_required 
def xoa_lop_hoc(request, lop_id):
    try:
        lop = get_object_or_404(LopHoc, id=lop_id)
        ten_lop = lop.ten_lop
        lop.delete()
        messages.success(request, f"Đã xóa lớp {ten_lop} thành công!")
    except Exception as e:
        messages.error(request, f"Lỗi khi xóa: {str(e)}")
    return redirect('quan_ly_lop_page')

@admin_required 
def dang_ky_lop(request, lop_id, hv_id):
    lop = get_object_or_404(LopHoc, id=lop_id)
    lop.hoc_viens.add(get_object_or_404(HocVien, id=hv_id))
    if hasattr(lop, 'cap_nhat_trang_thai'): lop.cap_nhat_trang_thai()
    return redirect('xep_lich_page')

@admin_required 
def xep_lich_page(request):
    if request.method == "POST":
        try:
            if GeneticScheduler().run():
                messages.success(request, "Thuật toán AI đã xếp lịch thành công!")
            else:
                messages.warning(request, "Chưa xếp được. Cần kiểm tra lại Lớp Học 'Đã Mở' hoặc số lượng Phòng.")
        except Exception as e:
            messages.error(request, f"Lỗi hệ thống thuật toán: {str(e)}")
        return redirect('xep_lich_page')

    # Dòng này phải thẳng hàng với chữ "if" ở trên (Tuyệt đối không thụt vào)
    danh_sach_co_so = CoSo.objects.all()
    lich_hoc_qs = LichHoc.objects.select_related('lop_hoc', 'phong', 'lop_hoc__giao_vien')
    
    # Logic lọc cơ sở an toàn tuyệt đối
    co_so_id = request.GET.get('co_so', '').strip()
    co_so_hien_tai = None
    
    if co_so_id and co_so_id.isdigit():
        co_so_hien_tai = int(co_so_id)
        lich_hoc_qs = lich_hoc_qs.filter(phong__co_so_id=co_so_hien_tai)
        
    lich_hoc = lich_hoc_qs.order_by('thu', 'ca')

    return render(request, 'xep_lich.html', {
        'lich_hoc': lich_hoc, 
        'danh_sach_co_so': danh_sach_co_so,
        'co_so_hien_tai': co_so_hien_tai
    })

@admin_required 
def tao_du_lieu_mau(request):
    tat_ca_co_so = CoSo.objects.all()
    if not tat_ca_co_so.exists():
        messages.error(request, "Cần tạo Cơ sở trước khi tạo dữ liệu mẫu!")
        return redirect('xep_lich_page')

    LichHoc.objects.all().delete()
    LopHoc.objects.all().delete()
    count_lop = 0
    hoc_viens = list(HocVien.objects.all())

    for cs in tat_ca_co_so:
        gv_toan, _ = GiaoVien.objects.get_or_create(ten=f"Thầy Toán ({cs.ten})", mon_day="Toán", co_so=cs)
        gv_ly, _ = GiaoVien.objects.get_or_create(ten=f"Cô Lý ({cs.ten})", mon_day="Lý", co_so=cs)
        gv_anh, _ = GiaoVien.objects.get_or_create(ten=f"Cô Anh ({cs.ten})", mon_day="Anh", co_so=cs)

        PhongHoc.objects.get_or_create(ten="P.101", suc_chua=30, co_so=cs)
        PhongHoc.objects.get_or_create(ten="P.102", suc_chua=30, co_so=cs)

        ds_lop = [("Toán 6A Nâng Cao", gv_toan), ("Toán 9C Ôn Thi", gv_toan), ("Lý 8B Cơ Bản", gv_ly), ("Anh 6C Giao Tiếp", gv_anh)]
        for ten_lop, gv in ds_lop:
            lop = LopHoc.objects.create(ten_lop=f"{ten_lop} - {cs.ten}", giao_vien=gv, min_si_so=5, trang_thai='MO')
            if len(hoc_viens) >= 5:
                lop.hoc_viens.set(random.sample(hoc_viens, random.randint(5, min(10, len(hoc_viens)))))
                lop.save()
            count_lop += 1

    messages.success(request, f"Đã reset và tạo tự động {count_lop} lớp học!")
    return redirect('xep_lich_page')

# ===================================================================
# 7. QUẢN LÝ TÀI CHÍNH
# ===================================================================
@admin_required 
def quan_ly_hoc_phi(request):
    if request.method == "POST":
        try:
            hv_id, so_tien, noi_dung, trang_thai = request.POST.get('hoc_vien'), request.POST.get('so_tien'), request.POST.get('noi_dung'), request.POST.get('trang_thai')
            if not hv_id or not so_tien:
                messages.error(request, "Vui lòng chọn học viên và nhập số tiền!")
            else:
                hv = HocVien.objects.get(id=hv_id)
                HoaDon.objects.create(hoc_vien=hv, so_tien=so_tien, noi_dung=noi_dung, trang_thai=trang_thai)
                messages.success(request, f"{'Đã thu tiền' if trang_thai == 'ROI' else 'Đã tạo công nợ'} cho: {hv.ten}")
        except Exception as e:
            messages.error(request, f"Lỗi: {e}")
        return redirect('quan_ly_hoc_phi')

    ds_hoa_don = HoaDon.objects.all().select_related('hoc_vien').order_by('-ngay_tao')
    stats = {
        'total_revenue': ds_hoa_don.filter(trang_thai='ROI').aggregate(Sum('so_tien'))['so_tien__sum'] or 0,
        'pending_debt': ds_hoa_don.filter(trang_thai='CHUA').aggregate(Sum('so_tien'))['so_tien__sum'] or 0,
        'total_invoices': ds_hoa_don.count(),
        'recent_paid': ds_hoa_don.filter(trang_thai='ROI').count()
    }
    return render(request, 'quan_ly_hoc_phi.html', {'ds_hoa_don': ds_hoa_don, 'ds_hoc_vien': HocVien.objects.all(), 'stats': stats})

@admin_required 
def cap_nhat_trang_thai(request, hd_id):
    """Nút duyệt thanh toán dành cho Admin"""
    try:
        hd = get_object_or_404(HoaDon, id=hd_id)
        # Sửa ở đây: Nếu đang CHUA hoặc CHO_DUYET -> Đổi thành ROI
        hd.trang_thai = 'ROI' if hd.trang_thai in ['CHUA', 'CHO_DUYET'] else 'CHUA'
        hd.save()
        messages.success(request, "Đã xác nhận thu tiền thành công!" if hd.trang_thai == 'ROI' else "Đã chuyển về trạng thái Chưa thanh toán.")
    except Exception as e:
        messages.error(request, f"Lỗi: {e}")
    return redirect('quan_ly_hoc_phi')

@admin_required
def xoa_hoa_don(request, hd_id):
    get_object_or_404(HoaDon, id=hd_id).delete()
    messages.success(request, "Đã hủy hóa đơn thành công!")
    return redirect('quan_ly_hoc_phi')

@admin_required
def xuat_excel_hoc_phi(request):
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse
    from django.utils import timezone
    from collections import defaultdict

    now = timezone.localtime(timezone.now())
    TRANG_THAI_MAP = {'ROI': 'Đã thanh toán', 'CHUA': 'Chưa thanh toán', 'CHO_DUYET': 'Chờ duyệt'}

    # --- Styles ---
    def _fill(hex_):  return PatternFill("solid", fgColor=hex_)
    def _align(h="center"): return Alignment(horizontal=h, vertical="center", wrap_text=True)
    _thin   = Side(style="thin",   color="BBBBBB")
    _medium = Side(style="medium", color="1E3A5F")
    border_data = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    border_title = Border(left=_medium, right=_medium, top=_medium, bottom=_medium)

    HEADERS    = ["STT", "Họ và Tên Học Viên", "Nội Dung", "Số Tiền (VNĐ)", "Trạng Thái", "Ngày Tạo"]
    COL_WIDTHS = [6, 30, 38, 20, 20, 22]
    NUM_COLS   = len(HEADERS)

    def _write_sheet_header(ws, tieu_de_thang):
        """Viết 4 dòng tiêu đề cho mỗi sheet."""
        # Dòng 1: Tên web
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NUM_COLS)
        c = ws.cell(row=1, column=1, value="TRUNG TÂM TOÁN GIS ENTERPRISE")
        c.font = Font(bold=True, color="FFFFFF", size=14, name="Calibri")
        c.fill = _fill("0D47A1")
        c.alignment = _align("center")
        c.border = border_title
        ws.row_dimensions[1].height = 30

        # Dòng 2: Tiêu đề tháng
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=NUM_COLS)
        c = ws.cell(row=2, column=1, value=f"BÁO CÁO HỌC PHÍ — {tieu_de_thang.upper()}")
        c.font = Font(bold=True, color="0D47A1", size=12, name="Calibri")
        c.fill = _fill("DBEAFE")
        c.alignment = _align("center")
        c.border = border_title
        ws.row_dimensions[2].height = 24

        # Dòng 3: Ngày xuất
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=NUM_COLS)
        c = ws.cell(row=3, column=1, value=f"Ngày xuất: {now.strftime('%d/%m/%Y %H:%M')}   |   Website: toan-gis.edu.vn")
        c.font = Font(italic=True, color="555555", size=9, name="Calibri")
        c.fill = _fill("F0F4FF")
        c.alignment = _align("center")
        ws.row_dimensions[3].height = 18

        # Dòng 4: Header cột
        for col, (h, w) in enumerate(zip(HEADERS, COL_WIDTHS), start=1):
            cell = ws.cell(row=4, column=col, value=h)
            cell.fill = _fill("1E3A5F")
            cell.font = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
            cell.alignment = _align("center")
            cell.border = border_data
            ws.column_dimensions[get_column_letter(col)].width = w
        ws.row_dimensions[4].height = 22
        ws.freeze_panes = "A5"

    def _write_data_rows(ws, ds_thang):
        """Ghi dữ liệu từ dòng 5 trở đi, trả về tổng tiền theo trạng thái."""
        tong = {'ROI': 0, 'CHUA': 0, 'CHO_DUYET': 0}
        for i, hd in enumerate(ds_thang, start=1):
            row = i + 4
            ws.row_dimensions[row].height = 18
            tt = hd.trang_thai
            tong[tt] = tong.get(tt, 0) + int(hd.so_tien)
            tt_display = TRANG_THAI_MAP.get(tt, tt)
            ngay = hd.ngay_tao.astimezone().strftime("%d/%m/%Y %H:%M") if hd.ngay_tao else ""

            mau_tt = _fill("D4EDDA") if tt == 'ROI' else (_fill("FFF3CD") if tt == 'CHO_DUYET' else _fill("F8D7DA"))
            row_data   = [i, hd.hoc_vien.ten, hd.noi_dung, int(hd.so_tien), tt_display, ngay]
            row_fills  = [_fill("DDE8F5"), _fill("E8F5E9"), _fill("FFF9E6"), _fill("FFF3CD"), mau_tt, _fill("EDE7F6")]
            row_aligns = ["center", "left", "left", "center", "center", "center"]
            row_bolds  = [True, True, False, False, False, False]

            for col, (val, fill, alg, bold) in enumerate(zip(row_data, row_fills, row_aligns, row_bolds), start=1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.fill = fill
                cell.font = Font(bold=bold, size=10, name="Calibri")
                cell.alignment = _align(alg)
                cell.border = border_data

        # Dòng tổng kết cuối sheet
        sum_row = len(ds_thang) + 5
        ws.merge_cells(start_row=sum_row, start_column=1, end_row=sum_row, end_column=3)
        ws.cell(row=sum_row, column=1, value="TỔNG KẾT THÁNG").font = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
        ws.cell(row=sum_row, column=1).fill = _fill("1E3A5F")
        ws.cell(row=sum_row, column=1).alignment = _align("center")

        tong_cell_data = [
            (4, f"{tong['ROI']:,} ₫", "D4EDDA", "155724"),
            (5, f"{tong['CHUA']:,} ₫", "F8D7DA", "721C24"),
            (6, f"{tong['CHO_DUYET']:,} ₫", "FFF3CD", "856404"),
        ]
        for col, val, bg, fg in tong_cell_data:
            c = ws.cell(row=sum_row, column=col, value=val)
            c.fill = _fill(bg); c.font = Font(bold=True, color=fg, size=10, name="Calibri")
            c.alignment = _align("center"); c.border = border_data
        ws.row_dimensions[sum_row].height = 22
        return tong

    # --- Nhóm hóa đơn theo tháng ---
    ds_all = HoaDon.objects.all().select_related('hoc_vien').order_by('ngay_tao')
    nhom_thang = defaultdict(list)
    for hd in ds_all:
        key = (hd.ngay_tao.year, hd.ngay_tao.month) if hd.ngay_tao else (0, 0)
        nhom_thang[key].append(hd)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Xóa sheet mặc định

    tong_toan_bo = {'ROI': 0, 'CHUA': 0, 'CHO_DUYET': 0}

    for (nam, thang) in sorted(nhom_thang.keys(), reverse=True):
        if thang == 0:
            ten_sheet = "Khong_Ngay"
            tieu_de = "Không Xác Định Tháng"
        else:
            ten_sheet = f"T{thang:02d}-{nam}"
            tieu_de = f"Tháng {thang}/{nam}"

        ws = wb.create_sheet(title=ten_sheet)
        _write_sheet_header(ws, tieu_de)
        tong = _write_data_rows(ws, nhom_thang[(nam, thang)])
        for k in tong_toan_bo:
            tong_toan_bo[k] += tong.get(k, 0)

    # --- Sheet tổng hợp ---
    ws_tong = wb.create_sheet(title="Tong_Hop", index=0)
    _write_sheet_header(ws_tong, f"Tổng Hợp Tất Cả Tháng")
    _write_data_rows(ws_tong, list(ds_all))

    ten_file = f"HocPhi_ToanGIS_{now.strftime('%m-%Y')}.xlsx"
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = f'attachment; filename="{ten_file}"'
    wb.save(response)
    return response

# ===================================================================
# 8. CỬA HÀNG KHÓA HỌC
# ===================================================================
from django.core.paginator import Paginator

@login_required(login_url='login')
def trang_khoa_hoc(request):
    if check_admin(request.user):
        messages.info(request, "Bạn là Quản trị viên, không cần đăng ký khóa học!")
        return redirect('thong_ke_page')

    # Lấy từ khóa tìm kiếm
    query = request.GET.get('q', '')

    # 1. Lọc và Phân trang cho Danh sách Khóa Học
    ds_khoa_hoc_qs = KhoaHoc.objects.all().order_by('-id')
    if query:
        ds_khoa_hoc_qs = ds_khoa_hoc_qs.filter(ten_khoa_hoc__icontains=query)
    
    paginator_kh = Paginator(ds_khoa_hoc_qs, 6) # Mỗi trang 6 khóa học
    page_kh = request.GET.get('page_kh', 1)
    danh_sach_khoa_hoc = paginator_kh.get_page(page_kh)

    # 2. Lọc và Phân trang cho Danh sách Lớp Học đang mở
    ds_lop_qs = LopHoc.objects.filter(trang_thai__in=['CHO', 'MO']).select_related('giao_vien').order_by('-id')
    if query:
        ds_lop_qs = ds_lop_qs.filter(ten_lop__icontains=query)
        
    paginator_lop = Paginator(ds_lop_qs, 6) # Mỗi trang 6 lớp học
    page_lop = request.GET.get('page_lop', 1)
    danh_sach_lop = paginator_lop.get_page(page_lop)

    return render(request, 'trang_khoa_hoc.html', {
        'danh_sach_khoa_hoc': danh_sach_khoa_hoc,
        'danh_sach_lop': danh_sach_lop,
        'query': query # Trả về từ khóa để giữ lại trên thanh tìm kiếm
    })

@login_required(login_url='login')
def user_dang_ky_lop(request, lop_id):
    """User mở trang xác nhận ghi danh và TỰ ĐỘNG SINH HÓA ĐƠN khi POST"""
    if check_admin(request.user):
        messages.warning(request, "Tài khoản Admin không thể tự đăng ký học!")
        return redirect('thong_ke_page') 
        
    lop = get_object_or_404(LopHoc, id=lop_id)
    hv, _ = HocVien.objects.get_or_create(ten=request.user.username, defaults={'khoi_lop': 12, 'lat': 0.0, 'lon': 0.0})

    if request.method == 'POST':
        try:
            if lop.hoc_viens.filter(id=hv.id).exists():
                messages.info(request, f"Bạn đã đăng ký lớp {lop.ten_lop} từ trước rồi!")
                return redirect('trang_khoa_hoc')
                
            elif lop.hoc_viens.count() >= lop.si_so_toi_da:
                messages.error(request, f"Rất tiếc, lớp {lop.ten_lop} đã đủ sĩ số tối đa!")
                return redirect('trang_khoa_hoc')
                
            else:
                # 1. Thêm học viên vào lớp
                lop.hoc_viens.add(hv)
                if hasattr(lop, 'cap_nhat_trang_thai'): 
                    lop.cap_nhat_trang_thai() 
                
                # 2. LOGIC TỰ ĐỘNG SINH HÓA ĐƠN (CÔNG NỢ)
                gia_tien = lop.khoa_hoc.hoc_phi if hasattr(lop, 'khoa_hoc') and lop.khoa_hoc else 1500000
                
                HoaDon.objects.create(
                    hoc_vien=hv,
                    so_tien=gia_tien,
                    noi_dung=f"Ghi danh: {lop.ten_lop}",
                    trang_thai='CHUA'
                )
                
                messages.success(request, f"Ghi danh thành công! Vui lòng thanh toán học phí để hoàn tất.")
                return redirect('trang_nguoi_dung') 
                
        except Exception as e:
            messages.error(request, f"Có lỗi xảy ra: {str(e)}")
            return redirect('trang_khoa_hoc')

    # Nếu là phương thức GET, render trang xác nhận đăng ký
    return render(request, 'user_dang_ky_lop.html', {'lop': lop})
# ===================================================================
# 9. QUẢN LÝ GIÁO VIÊN
# ===================================================================
@admin_required
def quan_ly_giao_vien(request):
    if request.method == "POST":
        gv_id = request.POST.get('gv_id')
        ten = request.POST.get('ten')
        tai_khoan = request.POST.get('tai_khoan')
        mon_day = request.POST.get('mon_day')
        co_so_id = request.POST.get('co_so')
        try:
            co_so = CoSo.objects.get(id=co_so_id) if co_so_id else None
            if gv_id: 
                gv = GiaoVien.objects.get(id=gv_id)
                gv.ten = ten; gv.tai_khoan = tai_khoan; gv.mon_day = mon_day; gv.co_so = co_so; gv.save()
                messages.success(request, f"Đã cập nhật hồ sơ giáo viên: {ten}")
            else:
                GiaoVien.objects.create(ten=ten, tai_khoan=tai_khoan, mon_day=mon_day, co_so=co_so)
                messages.success(request, f"Đã thêm giáo viên mới: {ten}")
        except Exception as e:
            messages.error(request, f"Lỗi hệ thống: {e}")
        return redirect('quan_ly_giao_vien')

    ds_giao_vien = GiaoVien.objects.all().select_related('co_so').order_by('-id')
    stats = {'total': ds_giao_vien.count(), 'toan': ds_giao_vien.filter(mon_day__icontains='Toán').count(), 'khac': ds_giao_vien.exclude(mon_day__icontains='Toán').count()}
    return render(request, 'quan_ly_giao_vien.html', {'ds_giao_vien': ds_giao_vien, 'ds_co_so': CoSo.objects.all(), 'stats': stats})

@admin_required
def xoa_giao_vien(request, gv_id):
    try:
        gv = get_object_or_404(GiaoVien, id=gv_id)
        ten = gv.ten; gv.delete()
        messages.success(request, f"Đã xóa giáo viên {ten} khỏi hệ thống.")
    except Exception as e:
        messages.error(request, f"Lỗi khi xóa: {str(e)}")
    return redirect('quan_ly_giao_vien')

# ===================================================================
# 10. QUẢN LÝ KHÓA HỌC
# ===================================================================
@admin_required
def quan_ly_khoa_hoc(request):
    if request.method == "POST":
        kh_id = request.POST.get('kh_id')
        ten = request.POST.get('ten_khoa_hoc')
        mo_ta = request.POST.get('mo_ta')
        hinh_anh = request.POST.get('hinh_anh')
        hoc_phi = request.POST.get('hoc_phi')
        try:
            if kh_id:
                kh = KhoaHoc.objects.get(id=kh_id)
                kh.ten_khoa_hoc = ten; kh.mo_ta = mo_ta; kh.hinh_anh = hinh_anh; kh.hoc_phi = hoc_phi; kh.save()
                messages.success(request, f"Đã cập nhật khóa học: {ten}")
            else:
                KhoaHoc.objects.create(ten_khoa_hoc=ten, mo_ta=mo_ta, hinh_anh=hinh_anh, hoc_phi=hoc_phi)
                messages.success(request, f"Đã khai trương khóa học mới: {ten}")
        except Exception as e:
            messages.error(request, f"Lỗi hệ thống: {e}")
        return redirect('quan_ly_khoa_hoc')
    ds_khoa_hoc = KhoaHoc.objects.all().order_by('-id')
    return render(request, 'quan_ly_khoa_hoc.html', {'ds_khoa_hoc': ds_khoa_hoc})

@admin_required
def xoa_khoa_hoc(request, kh_id):
    try:
        KhoaHoc.objects.get(id=kh_id).delete()
        messages.success(request, "Đã xóa khóa học thành công!")
    except Exception as e:
        messages.error(request, f"Lỗi: {e}")
    return redirect('quan_ly_khoa_hoc')

# ===================================================================
# 11. XỬ LÝ THANH TOÁN (HỌC VIÊN)
# ===================================================================
@login_required(login_url='login')
def thanh_toan_page(request):
    return render(request, 'thanh_toan.html', {})

@login_required(login_url='login')
def xu_ly_thanh_toan(request):
    """Xử lý nút Xác nhận quét QR và Nút gửi ảnh biên lai"""
    if request.method == 'POST':
        # 1. NẾU CÓ ĐÍNH KÈM FILE ẢNH (Gửi từ Bảng điều khiển)
        if request.FILES.get('anh_minh_chung') or request.FILES.getlist('anh_minh_chung'):
            danh_sach_anh = request.FILES.getlist('anh_minh_chung')
            hv = HocVien.objects.filter(ten=request.user.username).first()
            if hv and danh_sach_anh:
                # Tìm hóa đơn đang nợ để gắn ảnh vào
                hoa_don = HoaDon.objects.filter(hoc_vien=hv, trang_thai='CHUA').first()
                if hoa_don:
                    # Chạy vòng lặp lưu từng bức ảnh vào DB
                    for file_anh in danh_sach_anh:
                        AnhMinhChung.objects.create(hoa_don=hoa_don, hinh_anh=file_anh)
                    
                    hoa_don.trang_thai = 'CHO_DUYET' # Chuyển sang chờ duyệt
                    hoa_don.save()
                    messages.success(request, f"Đã gửi thành công {len(danh_sach_anh)} ảnh biên lai! Vui lòng chờ giáo vụ kiểm tra.")
                else:
                    messages.warning(request, "Bạn không có khoản nợ nào cần thanh toán.")
            return redirect('trang_nguoi_dung')
            
        # 2. NẾU CHỈ BẤM NÚT TỪ TRANG QUÉT QR (Chưa có ảnh)
        else:
            messages.info(request, "Vui lòng tải lên ảnh chụp màn hình biên lai tại đây để hoàn tất.")
            return redirect('trang_nguoi_dung')

    return redirect('trang_chu')

# [HÀM CHỐNG SẬP SERVER]
# Trong trường hợp Thầy quên xóa url upload_minh_chung trong urls.py
@login_required(login_url='login')
def upload_minh_chung(request, hd_id=None):
    return redirect('trang_nguoi_dung')

# ===================================================================
# 12. CÁC TÍNH NĂNG KHÁC (TÀI KHOẢN, ĐÁNH GIÁ, QUÊN MK)
# ===================================================================
@staff_member_required(login_url='login')
@staff_member_required(login_url='login')
def quan_ly_tai_khoan(request):
    # XỬ LÝ KHI ADMIN BẤM NÚT "XÁC NHẬN ĐỔI MẬT KHẨU"
    if request.method == "POST":
        user_id = request.POST.get('user_id')
        new_password = request.POST.get('new_password')
        
        if user_id and new_password:
            try:
                target_user = User.objects.get(id=user_id)
                # Dùng set_password để mã hóa mật khẩu an toàn theo chuẩn Django
                target_user.set_password(new_password)
                target_user.save()
                messages.success(request, f"Đã cấp lại mật khẩu cho tài khoản '{target_user.username}' thành công!")
            except Exception as e:
                messages.error(request, f"Lỗi hệ thống: {str(e)}")
        else:
            messages.error(request, "Vui lòng nhập mật khẩu mới!")
            
        return redirect('quan_ly_tai_khoan')

    # NẾU KHÔNG CÓ POST THÌ HIỂN THỊ DANH SÁCH BÌNH THƯỜNG
    danh_sach_user = User.objects.all().order_by('-date_joined')
    return render(request, 'quan_ly_tai_khoan.html', {'danh_sach_user': danh_sach_user})

def quen_mat_khau(request):
    if request.method == 'POST':
        thong_tin = request.POST.get('email_or_username', '').strip()
        user = User.objects.filter(email__iexact=thong_tin).first() or User.objects.filter(username__iexact=thong_tin).first()
        if user and user.email:
            mat_khau_moi = ''.join(random.choices(string.ascii_letters + string.digits, k=8))
            user.set_password(mat_khau_moi)
            user.save()
            tieu_de = 'Khôi phục mật khẩu - Hệ thống Toán GIS'
            noi_dung = f'''Chào {user.username},\nHệ thống vừa nhận được yêu cầu đặt lại mật khẩu của bạn.\nMật khẩu đăng nhập mới của bạn là: {mat_khau_moi}\nVui lòng đăng nhập lại và đổi mật khẩu ngay để bảo đảm an toàn.\nTrân trọng,\nBQT Toán GIS Enterprise.'''
            try:
                send_mail(tieu_de, noi_dung, settings.EMAIL_HOST_USER, [user.email])
                messages.success(request, f"Mật khẩu mới đã được gửi tới Email: {user.email}. Vui lòng kiểm tra hộp thư!")
            except Exception as e:
                messages.error(request, f"Hệ thống gặp lỗi khi kết nối Gmail: {str(e)}")
        else:
            messages.error(request, "Tài khoản không tồn tại hoặc chưa được liên kết Email!")
        return redirect('quen_mat_khau')
    return render(request, 'quen_mat_khau.html')

def danh_gia_page(request):
    if request.method == 'POST':
        try:
            # 1. KHÓA BẰNG TRÌNH DUYỆT (SESSION)
            if request.session.get('da_danh_gia', False):
                messages.error(request, "Bạn đã gửi đánh giá rồi! Mỗi người chỉ được tham gia 1 lần.")
                return redirect('danh_gia_page')

            # 2. KIỂM TRA CAPTCHA
            user_answer = request.POST.get('captcha_answer', '').strip()
            correct_answer = str(request.session.get('captcha_answer', ''))

            if not user_answer or user_answer != correct_answer:
                messages.error(request, "Sai mã xác thực! Vui lòng tính lại phép toán.")
                return redirect('danh_gia_page')

            # 3. LẤY DỮ LIỆU TỪ FORM
            ho_ten = request.POST.get('ho_ten', '').strip()
            email = request.POST.get('email', '').strip() # Lấy email
            co_so_id = request.POST.get('co_so')
            diem = int(request.POST.get('diem', 5))
            noi_dung = request.POST.get('noi_dung', '').strip()
            
            # SỬA Ở ĐÂY: BẮT BUỘC PHẢI CÓ EMAIL MỚI CHO QUA
            if not ho_ten or not email or not noi_dung:
                messages.error(request, "Vui lòng nhập đầy đủ Họ tên, Email và Nội dung đánh giá!")
            else:
                # 4. KHÓA BẰNG EMAIL DƯỚI DATABASE (Chống Tab Ẩn Danh)
                if DanhGia.objects.filter(email__iexact=email).exists():
                    messages.error(request, "Email này đã được sử dụng để đánh giá! Vui lòng dùng email khác.")
                    return redirect('danh_gia_page')

                # Lưu vào DB
                co_so = CoSo.objects.get(id=co_so_id) if co_so_id else None
                DanhGia.objects.create(ho_ten=ho_ten, email=email, co_so=co_so, diem=diem, noi_dung=noi_dung)
                
                # Khóa Session trình duyệt
                request.session['da_danh_gia'] = True
                messages.success(request, "Cảm ơn bạn đã gửi đánh giá! Ý kiến của bạn đã được ghi nhận.")
                
                if 'captcha_answer' in request.session:
                    del request.session['captcha_answer']
                    
        except Exception as e:
            messages.error(request, f"Lỗi: {e}")
        return redirect('danh_gia_page')

    # ... PHẦN GET BÊN DƯỚI GIỮ NGUYÊN ...
    num1 = random.randint(1, 10)
    num2 = random.randint(1, 10)
    request.session['captcha_answer'] = num1 + num2

    ds_danh_gia = DanhGia.objects.filter(duyet=True).select_related('co_so')
    ds_co_so = CoSo.objects.all()
    tong = ds_danh_gia.count()
    trung_binh = 0
    phan_bo = {5: 0, 4: 0, 3: 0, 2: 0, 1: 0}
    if tong > 0:
        tong_diem = sum(dg.diem for dg in ds_danh_gia)
        trung_binh = round(tong_diem / tong, 1)
        for dg in ds_danh_gia:
            phan_bo[dg.diem] = phan_bo.get(dg.diem, 0) + 1
    stats = {
        'tong': tong, 'trung_binh': trung_binh,
        'phan_bo': [(sao, sl, round(sl / tong * 100) if tong > 0 else 0) for sao, sl in sorted(phan_bo.items(), reverse=True)],
    }
    return render(request, 'danh_gia.html', {'ds_danh_gia': ds_danh_gia, 'ds_co_so': ds_co_so, 'stats': stats, 'num1': num1, 'num2': num2})

@admin_required
def xoa_danh_gia(request, dg_id):
    try:
        dg = get_object_or_404(DanhGia, id=dg_id)
        dg.delete()
        messages.success(request, "Đã xóa đánh giá!")
    except Exception as e:
        messages.error(request, f"Lỗi: {e}")
    return redirect('danh_gia_page')

# Trang danh sách tin tức cho người dùng (Sửa lỗi 500)
def trang_tin_tuc(request):
    # Phải có prefetch_related('cac_anh') để lấy danh sách ảnh đi kèm
    ds_tin_tuc = TinTuc.objects.prefetch_related('cac_anh').all().order_by('-ngay_dang')
    return render(request, 'tin_tuc.html', {'ds_tin_tuc': ds_tin_tuc})

@admin_required
def sua_tin_tuc(request, tin_id):
    tin = get_object_or_404(TinTuc, id=tin_id)
    if request.method == "POST":
        tin.tieu_de = request.POST.get('tieu_de')
        tin.loai_tin = request.POST.get('loai_tin')
        tin.tom_tat = request.POST.get('tom_tat')
        tin.noi_dung = request.POST.get('noi_dung')
        tin.save()
        
        # Thêm ảnh mới vào Album (nếu có)
        danh_sach_file = request.FILES.getlist('hinh_anh')
        for file in danh_sach_file:
            AnhTinTuc.objects.create(tin_tuc=tin, hinh_anh=file)
                
        messages.success(request, f"Đã cập nhật bài viết '{tin.tieu_de}' thành công!")
        return redirect('quan_ly_tin_tuc')
    
    return render(request, 'sua_tin_tuc.html', {'tin': tin})

@admin_required
def quan_ly_tin_tuc(request):
    if request.method == "POST":
        tieu_de = request.POST.get('tieu_de')
        tom_tat = request.POST.get('tom_tat')
        noi_dung = request.POST.get('noi_dung')
        loai_tin = request.POST.get('loai_tin')
        
        # 1. Tạo bài viết trước
        tin_moi = TinTuc.objects.create(
            tieu_de=tieu_de, 
            tom_tat=tom_tat, 
            noi_dung=noi_dung, 
            loai_tin=loai_tin
        )
        
        # 2. Lấy TẤT CẢ các ảnh được tải lên (dùng getlist) và lưu vào bảng AnhTinTuc
        danh_sach_file = request.FILES.getlist('hinh_anh')
        for file in danh_sach_file:
            AnhTinTuc.objects.create(tin_tuc=tin_moi, hinh_anh=file)
            
        messages.success(request, f"Đã đăng bài viết kèm {len(danh_sach_file)} hình ảnh thành công!")
        return redirect('quan_ly_tin_tuc')

    # Dùng prefetch_related để kéo luôn danh sách ảnh lên cho mượt
    ds_tin_tuc = TinTuc.objects.prefetch_related('cac_anh').order_by('-ngay_dang')
    return render(request, 'quan_ly_tin_tuc.html', {'ds_tin_tuc': ds_tin_tuc})
@admin_required
def xoa_tin_tuc(request, tin_id):
    # Lấy tin tức hoặc trả về 404 nếu không tồn tại
    tin = get_object_or_404(TinTuc, id=tin_id)
    
    # Lưu lại tiêu đề để hiện thông báo
    tieu_de = tin.tieu_de
    
    # Xóa tin tức (Django sẽ tự động xóa các AnhTinTuc liên quan nhờ on_delete=models.CASCADE)
    tin.delete()
    
    messages.success(request, f"Đã xóa bài viết '{tieu_de}' và toàn bộ ảnh liên quan thành công!")
    return redirect('quan_ly_tin_tuc')


# Nhớ import thêm model BaiGiang ở đầu file nhé: from .models import BaiGiang

@login_required(login_url='login')
def bai_giang_hoc_vien(request, lop_id):
    """Trang để học viên xem video và tải bài tập"""
    lop = get_object_or_404(LopHoc, id=lop_id)
    hv = HocVien.objects.filter(ten=request.user.username).first()
    
    # Bảo mật: Kiểm tra xem học viên có nằm trong lớp này không
    if not hv or not lop.hoc_viens.filter(id=hv.id).exists():
        messages.error(request, "Bạn chưa đăng ký hoặc chưa được duyệt vào lớp này!")
        return redirect('trang_nguoi_dung')
        
    ds_bai_giang = BaiGiang.objects.filter(lop_hoc=lop)
    return render(request, 'bai_giang.html', {'lop': lop, 'ds_bai_giang': ds_bai_giang})

@login_required(login_url='login')
def quan_ly_bai_giang_gv(request, lop_id):
    lop = get_object_or_404(LopHoc, id=lop_id)
    is_admin = check_admin(request.user)
    gv = GiaoVien.objects.filter(tai_khoan__iexact=request.user.username.strip()).first()
    
    if not is_admin and (not gv or lop.giao_vien != gv):
        messages.error(request, "Bạn không có quyền quản lý nội dung của lớp này!")
        return redirect('trang_giao_vien')
        
    if request.method == "POST":
        tieu_de = request.POST.get('tieu_de')
        mo_ta = request.POST.get('mo_ta')
        video_urls = request.POST.get('video_url', '') 
        
        # 1. Tạo bài giảng chính
        bai_moi = BaiGiang.objects.create(lop_hoc=lop, tieu_de=tieu_de, mo_ta=mo_ta)
        
        # 2a. XỬ LÝ LINK VIDEO YOUTUBE (Nếu có)
        links = [l.strip() for l in video_urls.replace(',', ' ').split() if l.strip()]
        for url in links:
            if "youtube.com/watch?v=" in url:
                url = url.replace("watch?v=", "embed/").replace("&", "?").split("?")[0]
            elif "youtu.be/" in url:
                vid_id = url.split("youtu.be/")[1].split("?")[0]
                url = f"https://www.youtube.com/embed/{vid_id}"
            VideoBaiGiang.objects.create(bai_giang=bai_moi, video_url=url)
            
        # 2b. XỬ LÝ FILE VIDEO TẢI LÊN (TÍNH NĂNG MỚI)
        video_files = request.FILES.getlist('video_files')
        for vf in video_files:
            VideoBaiGiang.objects.create(bai_giang=bai_moi, video_file=vf)
            
        # 3. Lưu danh sách nhiều file tài liệu
        files = request.FILES.getlist('file_bai_tap')
        for f in files:
            FileBaiGiang.objects.create(bai_giang=bai_moi, file_dinh_kem=f)
            
        messages.success(request, f"Đã đăng thành công {len(links)} link, {len(video_files)} video tải lên và {len(files)} tài liệu!")
        return redirect('quan_ly_bai_giang_gv', lop_id=lop.id)

    ds_bai_giang = BaiGiang.objects.filter(lop_hoc=lop).prefetch_related('danh_sach_video', 'danh_sach_file')
    return render(request, 'quan_ly_bai_giang.html', {'lop': lop, 'ds_bai_giang': ds_bai_giang})
@login_required(login_url='login')
def doi_mat_khau(request):
    if request.method == 'POST':
        # form = PasswordChangeForm(user, data)
        form = PasswordChangeForm(request.user, request.POST)
        if form.is_valid():
            user = form.save()
            # Giữ cho user không bị đăng xuất (văng ra ngoài) sau khi đổi MK thành công
            update_session_auth_hash(request, user)
            messages.success(request, 'Mật khẩu của bạn đã được cập nhật an toàn!')
            
            # Điều hướng về đúng trang của từng Role (Vai trò)
            if check_admin(request.user):
                return redirect('thong_ke_page')
            elif GiaoVien.objects.filter(tai_khoan__iexact=request.user.username.strip()).exists():
                return redirect('trang_giao_vien')
            else:
                return redirect('trang_nguoi_dung')
        else:
            # Nếu nhập sai (MK cũ sai, 2 MK mới không khớp, MK quá yếu...)
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, error)
    else:
        form = PasswordChangeForm(request.user)

    return render(request, 'doi_mat_khau.html')


# ===================================================================
# 13. BÀI TẬP VỀ NHÀ (BTVN)
# ===================================================================

@login_required(login_url='login')
def quan_ly_bai_tap_gv(request, lop_id):
    """GV tạo và quản lý bài tập cho lớp"""
    lop = get_object_or_404(LopHoc, id=lop_id)
    is_admin = check_admin(request.user)
    gv = GiaoVien.objects.filter(tai_khoan__iexact=request.user.username.strip()).first()

    if not is_admin and (not gv or lop.giao_vien != gv):
        messages.error(request, "Bạn không có quyền quản lý lớp này!")
        return redirect('trang_giao_vien')

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'delete':
            bt_id = request.POST.get('bt_id')
            get_object_or_404(BaiTap, id=bt_id, lop_hoc=lop).delete()
            messages.success(request, "Đã xóa bài tập!")
            return redirect('quan_ly_bai_tap_gv', lop_id=lop.id)

        tieu_de = request.POST.get('tieu_de', '').strip()
        if not tieu_de:
            messages.error(request, "Vui lòng nhập tiêu đề bài tập!")
        else:
            han_nop_str = request.POST.get('han_nop', '')
            bt = BaiTap.objects.create(
                lop_hoc=lop,
                tieu_de=tieu_de,
                mo_ta=request.POST.get('mo_ta', ''),
                hinh_thuc=request.POST.get('hinh_thuc', 'ONL'),
                han_nop=han_nop_str if han_nop_str else None,
            )
            if request.FILES.get('file_de'):
                bt.file_de = request.FILES['file_de']
                bt.save()
            messages.success(request, f"Đã giao bài tập: {tieu_de}")
        return redirect('quan_ly_bai_tap_gv', lop_id=lop.id)

    ds_bai_tap = BaiTap.objects.filter(lop_hoc=lop).prefetch_related('bai_nop')
    return render(request, 'quan_ly_bai_tap_gv.html', {'lop': lop, 'ds_bai_tap': ds_bai_tap})


@login_required(login_url='login')
def bai_tap_hoc_vien(request, lop_id):
    """HV xem danh sách bài tập của lớp"""
    lop = get_object_or_404(LopHoc, id=lop_id)
    hv = HocVien.objects.filter(ten=request.user.username).first()

    if not hv or not lop.hoc_viens.filter(id=hv.id).exists():
        messages.error(request, "Bạn chưa đăng ký lớp này!")
        return redirect('trang_nguoi_dung')

    ds_bai_tap = BaiTap.objects.filter(lop_hoc=lop).order_by('-ngay_tao')
    bai_da_nop_map = {n.bai_tap_id: n for n in NopBaiTap.objects.filter(hoc_vien=hv, bai_tap__lop_hoc=lop)}
    ds_bai_tap_voi_nop = [(bt, bai_da_nop_map.get(bt.id)) for bt in ds_bai_tap]
    return render(request, 'bai_tap_hoc_vien.html', {
        'lop': lop, 'ds_bai_tap_voi_nop': ds_bai_tap_voi_nop
    })


@login_required(login_url='login')
def nop_bai_tap(request, bt_id):
    """HV nộp bài tập"""
    bt = get_object_or_404(BaiTap, id=bt_id)
    hv = HocVien.objects.filter(ten=request.user.username).first()

    if not hv or not bt.lop_hoc.hoc_viens.filter(id=hv.id).exists():
        messages.error(request, "Bạn không có quyền nộp bài này!")
        return redirect('trang_nguoi_dung')

    da_nop = NopBaiTap.objects.filter(bai_tap=bt, hoc_vien=hv).first()

    if request.method == 'POST':
        ghi_chu = request.POST.get('ghi_chu', '')
        if da_nop:
            da_nop.ghi_chu = ghi_chu
            if request.FILES.get('file_nop'):
                da_nop.file_nop = request.FILES['file_nop']
            da_nop.save()
            messages.success(request, "Đã cập nhật bài nộp!")
        else:
            nop = NopBaiTap(bai_tap=bt, hoc_vien=hv, ghi_chu=ghi_chu)
            if request.FILES.get('file_nop'):
                nop.file_nop = request.FILES['file_nop']
            nop.save()
            messages.success(request, "Nộp bài thành công!")
        return redirect('bai_tap_hoc_vien', lop_id=bt.lop_hoc.id)

    return render(request, 'nop_bai_tap.html', {'bt': bt, 'da_nop': da_nop})


@login_required(login_url='login')
def cham_bai_tap(request, bt_id):
    """GV xem danh sách bài nộp và chấm điểm"""
    bt = get_object_or_404(BaiTap, id=bt_id)
    is_admin = check_admin(request.user)
    gv = GiaoVien.objects.filter(tai_khoan__iexact=request.user.username.strip()).first()

    if not is_admin and (not gv or bt.lop_hoc.giao_vien != gv):
        messages.error(request, "Bạn không có quyền chấm bài này!")
        return redirect('trang_giao_vien')

    if request.method == 'POST':
        nop_id = request.POST.get('nop_id')
        nop = get_object_or_404(NopBaiTap, id=nop_id)
        diem_str = request.POST.get('diem', '').strip()
        nop.diem = float(diem_str) if diem_str else None
        nop.nhan_xet = request.POST.get('nhan_xet', '')
        nop.save()
        messages.success(request, f"Đã chấm điểm cho {nop.hoc_vien.ten}!")
        return redirect('cham_bai_tap', bt_id=bt.id)

    ds_nop = NopBaiTap.objects.filter(bai_tap=bt).select_related('hoc_vien')
    da_nop_ids = set(ds_nop.values_list('hoc_vien_id', flat=True))
    chua_nop = bt.lop_hoc.hoc_viens.exclude(id__in=da_nop_ids)
    return render(request, 'cham_bai_tap.html', {'bt': bt, 'ds_nop': ds_nop, 'chua_nop': chua_nop})


# ===================================================================
# 14. THỜI GIAN RẢNH CỦA GIÁO VIÊN
# ===================================================================

@login_required(login_url='login')
def thoi_gian_ranh_gv(request):
    """GV đăng ký ca học rảnh / xóa ca bận"""
    username_sach = request.user.username.strip()
    gv = GiaoVien.objects.filter(tai_khoan__iexact=username_sach).first()
    if not gv and not check_admin(request.user):
        messages.error(request, "Chỉ giáo viên mới có thể đăng ký thời gian rảnh!")
        return redirect('trang_nguoi_dung')

    if gv is None:
        messages.info(request, "Admin: chọn giáo viên từ trang quản lý.")
        return redirect('quan_ly_giao_vien')

    if request.method == 'POST':
        action = request.POST.get('action')
        thu = request.POST.get('thu')
        ca = request.POST.get('ca')
        if action == 'xoa':
            ThoiGianRanh.objects.filter(giao_vien=gv, thu=thu, ca=ca).delete()
            messages.success(request, f"Đã xóa ca rảnh Thứ {thu} Ca {ca}.")
        else:
            ThoiGianRanh.objects.get_or_create(giao_vien=gv, thu=thu, ca=ca)
            messages.success(request, f"Đã thêm ca rảnh Thứ {thu} Ca {ca}.")
        return redirect('thoi_gian_ranh_gv')

    ds_ranh = ThoiGianRanh.objects.filter(giao_vien=gv)
    ranh_set = {(r.thu, r.ca) for r in ds_ranh}
    grid = [
        {'ca': ca, 'cells': [{'thu': thu, 'ca': ca, 'is_ranh': (thu, ca) in ranh_set} for thu in range(2, 9)]}
        for ca in range(1, 7)
    ]
    return render(request, 'thoi_gian_ranh.html', {
        'giao_vien': gv, 'ds_ranh': ds_ranh, 'grid': grid
    })


@admin_required
def xem_thoi_gian_ranh_admin(request):
    """Admin xem lịch rảnh của tất cả GV"""
    ds_gv = GiaoVien.objects.prefetch_related('thoi_gian_ranh').all()
    return render(request, 'thoi_gian_ranh_admin.html', {
        'ds_gv': ds_gv, 'thu_range': range(2, 9), 'ca_range': range(1, 7)
    })


# ===================================================================
# 15. XIN NGHỈ & LỊCH BÙ
# ===================================================================

@login_required(login_url='login')
def xin_nghi(request, lop_id):
    """HV hoặc GV gửi đơn xin nghỉ một buổi"""
    lop = get_object_or_404(LopHoc, id=lop_id)
    username_sach = request.user.username.strip()
    hv = HocVien.objects.filter(ten=username_sach).first()
    gv = GiaoVien.objects.filter(tai_khoan__iexact=username_sach).first()

    if not hv and not gv and not check_admin(request.user):
        messages.error(request, "Không tìm thấy thông tin tài khoản!")
        return redirect('trang_nguoi_dung')

    if request.method == 'POST':
        ngay_nghi = request.POST.get('ngay_nghi')
        ly_do = request.POST.get('ly_do', '').strip()
        if not ngay_nghi or not ly_do:
            messages.error(request, "Vui lòng điền đầy đủ ngày nghỉ và lý do!")
        else:
            loai = 'GV' if gv else 'HV'
            XinNghi.objects.create(
                loai=loai,
                hoc_vien=hv if loai == 'HV' else None,
                giao_vien=gv if loai == 'GV' else None,
                lop_hoc=lop,
                ngay_nghi=ngay_nghi,
                ly_do=ly_do,
            )
            messages.success(request, "Đã gửi đơn xin nghỉ, chờ admin duyệt.")
            return redirect('trang_giao_vien' if gv else 'trang_nguoi_dung')

    return render(request, 'xin_nghi.html', {'lop': lop, 'hoc_vien': hv, 'giao_vien': gv})


@admin_required
def quan_ly_xin_nghi(request):
    """Admin duyệt / từ chối đơn xin nghỉ"""
    if request.method == 'POST':
        xn_id = request.POST.get('xn_id')
        action = request.POST.get('action')
        xn = get_object_or_404(XinNghi, id=xn_id)
        if action == 'duyet':
            xn.trang_thai = 'DUYET'
            xn.save()
            messages.success(request, f"Đã duyệt đơn xin nghỉ của {xn.ten_nguoi_xin()}.")
        elif action == 'tu_choi':
            xn.trang_thai = 'TU_CHOI'
            xn.save()
            messages.success(request, f"Đã từ chối đơn của {xn.ten_nguoi_xin()}.")
        return redirect('quan_ly_xin_nghi')

    ds_xin_nghi = XinNghi.objects.select_related('hoc_vien', 'giao_vien', 'lop_hoc').order_by('-ngay_tao')
    filter_trang_thai = request.GET.get('tt', '')
    if filter_trang_thai:
        ds_xin_nghi = ds_xin_nghi.filter(trang_thai=filter_trang_thai)

    stats = {
        'cho': XinNghi.objects.filter(trang_thai='CHO').count(),
        'duyet': XinNghi.objects.filter(trang_thai='DUYET').count(),
        'tu_choi': XinNghi.objects.filter(trang_thai='TU_CHOI').count(),
    }
    return render(request, 'quan_ly_xin_nghi.html', {
        'ds_xin_nghi': ds_xin_nghi, 'stats': stats, 'filter_tt': filter_trang_thai
    })


@login_required(login_url='login')
def tao_lich_bu(request):
    """GV hoặc Admin tạo lịch học bù và thông báo cho HV"""
    username_sach = request.user.username.strip()
    gv = GiaoVien.objects.filter(tai_khoan__iexact=username_sach).first()
    is_admin = check_admin(request.user)

    if not gv and not is_admin:
        messages.error(request, "Chỉ giáo viên hoặc admin mới có thể tạo lịch bù!")
        return redirect('trang_nguoi_dung')

    if request.method == 'POST':
        lop_id = request.POST.get('lop_hoc')
        ngay_bu = request.POST.get('ngay_bu')
        ca_bu = request.POST.get('ca_bu')
        ly_do = request.POST.get('ly_do', '')
        xn_id = request.POST.get('xin_nghi_id', '')

        if not lop_id or not ngay_bu or not ca_bu:
            messages.error(request, "Vui lòng điền đầy đủ thông tin lịch bù!")
        else:
            lop = get_object_or_404(LopHoc, id=lop_id)
            xn = XinNghi.objects.filter(id=xn_id).first() if xn_id else None
            lb = LichBu.objects.create(
                lop_hoc=lop,
                xin_nghi=xn,
                ngay_bu=ngay_bu,
                ca_bu=ca_bu,
                ly_do=ly_do,
                trang_thai='DUYET',
            )
            messages.success(request, f"Đã tạo lịch bù cho lớp {lop.ten_lop} ngày {ngay_bu}!")
            return redirect('xem_lich_bu')

    ds_lop = LopHoc.objects.filter(giao_vien=gv) if gv else LopHoc.objects.all()
    ds_xin_nghi_duyet = XinNghi.objects.filter(trang_thai='DUYET').select_related('lop_hoc', 'hoc_vien', 'giao_vien')
    return render(request, 'tao_lich_bu.html', {
        'ds_lop': ds_lop, 'ds_xin_nghi': ds_xin_nghi_duyet, 'giao_vien': gv,
        'ca_range': range(1, 7)
    })


@login_required(login_url='login')
def xem_lich_bu(request):
    """Trang danh sách lịch bù — GV, HV và Admin đều xem được"""
    username_sach = request.user.username.strip()
    is_admin = check_admin(request.user)
    gv = GiaoVien.objects.filter(tai_khoan__iexact=username_sach).first()
    hv = HocVien.objects.filter(ten=username_sach).first()

    if is_admin:
        ds_lb = LichBu.objects.select_related('lop_hoc', 'xin_nghi').prefetch_related('danh_sach_dang_ky').order_by('-ngay_bu')
    elif gv:
        ds_lb = LichBu.objects.filter(lop_hoc__giao_vien=gv).select_related('lop_hoc').prefetch_related('danh_sach_dang_ky').order_by('-ngay_bu')
    elif hv:
        ds_lb = LichBu.objects.filter(lop_hoc__hoc_viens=hv).select_related('lop_hoc').prefetch_related('danh_sach_dang_ky').order_by('-ngay_bu')
    else:
        ds_lb = LichBu.objects.none()

    da_dang_ky = {}
    da_dang_ky_ids = set()
    if hv:
        for dk in DangKyLichBu.objects.filter(hoc_vien=hv):
            da_dang_ky[dk.lich_bu_id] = dk
            da_dang_ky_ids.add(dk.lich_bu_id)

    return render(request, 'xem_lich_bu.html', {
        'ds_lb': ds_lb, 'is_admin': is_admin, 'giao_vien': gv,
        'hoc_vien': hv, 'da_dang_ky': da_dang_ky, 'da_dang_ky_ids': da_dang_ky_ids
    })


@login_required(login_url='login')
def dang_ky_lich_bu(request, lb_id):
    """HV xác nhận có thể tham dự hay không buổi bù"""
    lb = get_object_or_404(LichBu, id=lb_id)
    hv = HocVien.objects.filter(ten=request.user.username).first()

    if not hv or not lb.lop_hoc.hoc_viens.filter(id=hv.id).exists():
        messages.error(request, "Bạn không thuộc lớp học này!")
        return redirect('xem_lich_bu')

    if request.method == 'POST':
        co_mat = request.POST.get('co_mat') == '1'
        ghi_chu = request.POST.get('ghi_chu', '')
        dk, created = DangKyLichBu.objects.update_or_create(
            hoc_vien=hv, lich_bu=lb,
            defaults={'co_mat': co_mat, 'ghi_chu': ghi_chu}
        )
        trang_thai_text = "tham dự" if co_mat else "vắng mặt"
        messages.success(request, f"Đã ghi nhận: bạn sẽ {trang_thai_text} buổi bù ngày {lb.ngay_bu}.")
        return redirect('xem_lich_bu')

    da_dk = DangKyLichBu.objects.filter(hoc_vien=hv, lich_bu=lb).first()
    return render(request, 'dang_ky_lich_bu.html', {'lb': lb, 'da_dk': da_dk})


# ===================================================================
# GIỚI THIỆU TRUNG TÂM
# ===================================================================

def trang_gioi_thieu(request):
    ds_gioi_thieu = GioiThieu.objects.prefetch_related('cac_anh').all()
    return render(request, 'gioi_thieu.html', {'ds_gioi_thieu': ds_gioi_thieu})


@admin_required
def quan_ly_gioi_thieu(request):
    if request.method == 'POST':
        tieu_de = request.POST.get('tieu_de')
        tom_tat = request.POST.get('tom_tat')
        noi_dung = request.POST.get('noi_dung')
        loai_muc = request.POST.get('loai_muc')
        thu_tu = request.POST.get('thu_tu', 0)

        muc_moi = GioiThieu.objects.create(
            tieu_de=tieu_de,
            tom_tat=tom_tat,
            noi_dung=noi_dung,
            loai_muc=loai_muc,
            thu_tu=int(thu_tu),
        )

        danh_sach_file = request.FILES.getlist('hinh_anh')
        for file in danh_sach_file:
            AnhGioiThieu.objects.create(gioi_thieu=muc_moi, hinh_anh=file)

        messages.success(request, f"Đã thêm mục giới thiệu '{tieu_de}' với {len(danh_sach_file)} hình ảnh!")
        return redirect('quan_ly_gioi_thieu')

    ds_gioi_thieu = GioiThieu.objects.prefetch_related('cac_anh').all()
    return render(request, 'quan_ly_gioi_thieu.html', {'ds_gioi_thieu': ds_gioi_thieu})


@admin_required
def sua_gioi_thieu(request, gt_id):
    muc = get_object_or_404(GioiThieu, id=gt_id)
    if request.method == 'POST':
        muc.tieu_de = request.POST.get('tieu_de')
        muc.tom_tat = request.POST.get('tom_tat')
        muc.noi_dung = request.POST.get('noi_dung')
        muc.loai_muc = request.POST.get('loai_muc')
        muc.thu_tu = int(request.POST.get('thu_tu', 0))
        muc.save()

        danh_sach_file = request.FILES.getlist('hinh_anh')
        for file in danh_sach_file:
            AnhGioiThieu.objects.create(gioi_thieu=muc, hinh_anh=file)

        messages.success(request, "Đã cập nhật mục giới thiệu thành công!")
        return redirect('quan_ly_gioi_thieu')

    return render(request, 'sua_gioi_thieu.html', {'muc': muc})


@admin_required
def xoa_gioi_thieu(request, gt_id):
    muc = get_object_or_404(GioiThieu, id=gt_id)
    tieu_de = muc.tieu_de
    muc.delete()
    messages.success(request, f"Đã xóa mục '{tieu_de}' và toàn bộ hình ảnh liên quan!")
    return redirect('quan_ly_gioi_thieu')


@admin_required
def xoa_anh_gioi_thieu(request, anh_id):
    anh = get_object_or_404(AnhGioiThieu, id=anh_id)
    gt_id = anh.gioi_thieu.id
    anh.hinh_anh.delete(save=False)
    anh.delete()
    messages.success(request, "Đã xóa ảnh!")
    return redirect('sua_gioi_thieu', gt_id=gt_id)